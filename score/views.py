from django.shortcuts import render

# Create your views here.
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from .forms import EvaluationForm
from .models import Evaluation, Student
from django.contrib import messages
from django.http import HttpResponse
import csv
import openpyxl
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.forms import AuthenticationForm
from .forms import EvaluatorRegistrationForm
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
def is_coordinator(user):
    return user.is_superuser  # You can customize this check

@login_required
def submit_score(request):
    if request.method == 'POST':
        form = EvaluationForm(request.POST)
        if form.is_valid():
            student = form.cleaned_data['student']
            existing_eval = Evaluation.objects.filter(evaluator=request.user, student=student).first()
            if existing_eval:
                messages.error(request, 'You have already scored this student.')
            elif Evaluation.objects.filter(student=student).count() >= 4:
                messages.error(request, 'This student has already been evaluated by 4 evaluators.')
            else:
                evaluation = form.save(commit=False)
                evaluation.evaluator = request.user
                evaluation.save()
                messages.success(request, 'Score submitted successfully.')
                return redirect('submit_score')
    else:
        form = EvaluationForm()
    return render(request, 'submit_score.html', {'form': form})


@login_required
@user_passes_test(is_coordinator)
def coordinator_dashboard(request):
    student_list = Student.objects.all()
    results = []

    for student in student_list:
        evaluations = Evaluation.objects.filter(student=student)
        evaluator_count = evaluations.count()
        evaluators_remaining = 4 - evaluator_count

        if evaluator_count == 4:
            total_score = sum(e.total_score for e in evaluations)
            average_score = round(total_score / 4, 2)
            score_display = f"{average_score} / 100"
        else:
            score_display = "Awaiting Result"

        results.append({
            'matric_number': student.matric_number,
            'full_name': student.full_name,
            'project_topic': student.project_topic,
            'department': student.department,
            'total_score': score_display,
            'remaining': evaluators_remaining
        })

    paginator = Paginator(results, 10)  # Show 10 students per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'dashboard.html', {'page_obj': page_obj})

@login_required
@user_passes_test(is_coordinator)
def download_results_csv(request):
    students = Student.objects.all()
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="project_scores.csv"'

    writer = csv.writer(response)
    writer.writerow(['Matric Number', 'Project Topic', 'Department', 'Total Score'])

    for student in students:
        evaluations = Evaluation.objects.filter(student=student)
        if evaluations.count() == 4:
            total_score = sum(e.total_score for e in evaluations)
            average_score = round(total_score / 4, 2)
            writer.writerow([student.matric_number, student.project_topic, student.department, average_score])
        else:
            writer.writerow([student.matric_number, student.project_topic, student.department, "Awaiting Result"])

    return response

@login_required
@user_passes_test(is_coordinator)
def import_students_from_excel(request):
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']

        if not file.name.endswith('.xlsx'):
            messages.error(request, 'Invalid file format. Please upload a .xlsx Excel file.')
            return redirect('import_students')

        try:
            wb = openpyxl.load_workbook(file)
            sheet = wb.active
            expected_columns = ['Full Name', 'Matric Number', 'Project Topic', 'Department']
            header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

            # Validate header structure
            if header != expected_columns:
                messages.error(request, f"Invalid file structure. Expected columns: {', '.join(expected_columns)}")
                return redirect('import_students')

            imported_count = 0
            skipped = []

            for row in sheet.iter_rows(min_row=2, values_only=True):
                full_name, matric_number, project_topic, department = row

                if not all([full_name, matric_number, project_topic, department]):
                    skipped.append(matric_number or 'Unknown')
                    continue

                _, created = Student.objects.get_or_create(
                    matric_number=matric_number,
                    defaults={
                        'full_name': full_name,
                        'project_topic': project_topic,
                        'department': department
                    }
                )
                if created:
                    imported_count += 1
                else:
                    skipped.append(matric_number)

            messages.success(request, f'{imported_count} students imported successfully.')
            if skipped:
                messages.warning(request, f'Skipped {len(skipped)} existing or invalid students: {", ".join(skipped[:5])}{"..." if len(skipped) > 5 else ""}')

        except Exception as e:
            messages.error(request, f'Error processing file: {str(e)}')

        return redirect('import_students')

    return render(request, 'import_students.html')



def register_view(request):
    if request.method == 'POST':
        form = EvaluatorRegistrationForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Registration successful. You can now log in.")
            return redirect('login_view')
    else:
        form = EvaluatorRegistrationForm()
    return render(request, 'register.html', {'form': form})

def login_view(request):
    if request.user.is_authenticated:
        return redirect('submit_score')
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect('submit_score')
        messages.error(request, "Invalid username or password.")
    else:
        form = AuthenticationForm()
    return render(request, 'login.html', {'form': form})


@login_required
def logout_view(request):
    logout(request)
    messages.info(request, "Logged out successfully.")
    return redirect('login_view')


def check_grade(request):
    result = None
    error = None

    if request.method == 'POST':
        matric_number = request.POST.get('matric_number').strip().upper()
        try:
            student = Student.objects.get(matric_number=matric_number)
            evaluations = Evaluation.objects.filter(student=student)
            evaluator_count = evaluations.count()

            if evaluator_count == 4:
                total_score = sum(e.total_score for e in evaluations)
                average_score = round(total_score / 4, 2)
                score_display = f"{average_score} / 50"
            else:
                score_display = "Awaiting Result"

            result = {
                'full_name': student.full_name,
                'matric_number': student.matric_number,
                'project_topic': student.project_topic,
                'department': student.department,
                'total_score': score_display
            }
        except Student.DoesNotExist:
            error = "Matric number not found. Please check and try again."

    return render(request, 'check_grade.html', {'result': result, 'error': error})
